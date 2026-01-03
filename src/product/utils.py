import math
import os
import secrets
from io import BytesIO
from typing import List

import openpyxl
import qrcode
import requests
from cloudinary.uploader import upload
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

from account.exceptions import BackendURLNotConfigured
from product.models import Product, VerifyCode


PRODUCT_HEADERS = [
    "Tên sản phẩm",
    "Mã sản phẩm",
    "Giá gốc",
    "Giá bán",
    "Thương hiệu",
    "Phân loại",
    "Mô tả",
    "Số lượng trong kho",
    "Ảnh đại diện",
]


def build_product_workbook() -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách sản phẩm"

    ws.append(PRODUCT_HEADERS)
    for index in range(1, len(PRODUCT_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(index)].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def extract_images(sheet):
    image_map = {}
    for img in sheet._images:
        anchor = img.anchor
        if hasattr(anchor, "_from"):
            row = anchor._from.row + 1
        else:
            row = anchor.from_.row + 1
        image_map[row] = img
    return image_map


def load_product_infomation(product_file) -> List[dict]:
    workbook = openpyxl.load_workbook(product_file)
    sheet = workbook.active
    image_map = extract_images(sheet)
    products = []

    for row_index, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        (
            name,
            code,
            origin_price,
            sale_price,
            brand_name,
            type_,
            description,
            quantity_in_stock,
        ) = row

        if not all([name, code, origin_price, sale_price, type_, quantity_in_stock]):
            continue

        product_data = {
            "name": name,
            "code": code,
            "origin_price": origin_price,
            "sale_price": sale_price,
            "brand_name": brand_name,
            "type": type_,
            "description": description or "",
            "quantity_in_stock": quantity_in_stock,
            "image": image_map.get(row_index),
        }

        products.append(product_data)

    return products


def upload_file(file, folder: str, public_id: str, overwrite: bool = True) -> dict:
    """
    Upload 1 file lên cloud storage (Cloudinary).
    Trả về dict kết quả chứa URL, public_id, ...
    """
    return upload(
        file,
        folder=folder,
        public_id=public_id,
        overwrite=overwrite,
    )


def generate_qrcode(product: Product):
    code = secrets.token_urlsafe(32)
    backend_url = os.environ.get("BACKEND_URL")
    if not backend_url:
        raise BackendURLNotConfigured
    link = f"{backend_url}/api/verifycodes/verify-qrcode?code={code}"
    qr = qrcode.make(link)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    buffer.seek(0)
    result = upload_file(file=buffer, folder="qr_codes/", public_id=f"verify_{code}")

    verify_code = VerifyCode.objects.create(
        product=product,
        code=code,
        qr_url=result["secure_url"],
    )

    return verify_code


def generate_qrcode_pdf(verify_codes):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    cols = 10
    gap = 0.5 * mm
    padding = 5 * mm

    usable_width = width - 2 * padding
    usable_height = height - 2 * padding

    size = (usable_width - (cols - 1) * gap) / cols

    rows = math.floor((usable_height + gap) / (size + gap))

    x_start = padding
    y_start = height - padding - size

    x = x_start
    y = y_start
    count = 0

    for vc in verify_codes:
        if not vc.qr_url:
            continue

        r = requests.get(vc.qr_url, timeout=10)
        if r.status_code != 200:
            continue

        if not r.headers.get("Content-Type", "").startswith("image/"):
            continue

        image = ImageReader(BytesIO(r.content))
        c.drawImage(image, x, y, size, size, preserveAspectRatio=True)

        x += size + gap
        count += 1

        if count % cols == 0:
            x = x_start
            y -= size + gap

        if count % (cols * rows) == 0:
            c.showPage()
            x = x_start
            y = y_start

    c.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="qr_codes.pdf"'
    return response


VERIFY_QR_TEMPLATE = """
<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Xác minh sản phẩm</title>

    <style>
        * {
            box-sizing: border-box;
        }

        body {
            margin: 0;
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            background: linear-gradient(135deg, #f8fafc, #eef2ff);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 16px;
        }

        .card {
            width: 100%;
            max-width: 420px;
            background: #ffffff;
            border-radius: 20px;
            padding: 24px;
            box-shadow: 0 20px 40px rgba(0,0,0,.08);
            animation: fadeIn .4s ease;
        }

        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(8px); }
            to { opacity: 1; transform: translateY(0); }
        }

        .status {
            text-align: center;
            font-size: 22px;
            font-weight: 800;
            margin-bottom: 8px;
            letter-spacing: .5px;
        }

        .AUTHENTIC { color: #16a34a; }
        .SCANNED   { color: #d97706; }
        .FAKE      { color: #dc2626; }

        .icon {
            font-size: 48px;
            text-align: center;
            margin-bottom: 12px;
        }

        .message {
            text-align: center;
            color: #374151;
            margin-bottom: 20px;
            font-size: 15px;
        }

        .divider {
            height: 1px;
            background: #e5e7eb;
            margin: 20px 0;
        }

        .product img {
            width: 100%;
            border-radius: 14px;
            margin-bottom: 14px;
            object-fit: cover;
        }

        .product h3 {
            margin: 0 0 6px;
            font-size: 18px;
        }

        .meta {
            font-size: 14px;
            color: #6b7280;
            margin-bottom: 4px;
        }

        .badge {
            margin-top: 14px;
            display: inline-block;
            padding: 8px 14px;
            border-radius: 999px;
            font-size: 13px;
            font-weight: 600;
        }

        .green {
            background: #dcfce7;
            color: #166534;
        }

        .yellow {
            background: #fef3c7;
            color: #92400e;
        }

        .red {
            background: #fee2e2;
            color: #991b1b;
        }

        .footer {
            text-align: center;
            font-size: 12px;
            color: #9ca3af;
            margin-top: 24px;
        }
    </style>
</head>

<body>

<div class="card">

    <!-- ICON -->
    <div class="icon">
        {% if status == "AUTHENTIC" %}
            ✅
        {% elif status == "SCANNED" %}
            ⚠️
        {% else %}
            ❌
        {% endif %}
    </div>

    <!-- STATUS -->
    <div class="status {{ status }}">
        {% if status == "AUTHENTIC" %}
            SẢN PHẨM CHÍNH HÃNG
        {% elif status == "SCANNED" %}
            MÃ ĐÃ ĐƯỢC QUÉT
        {% else %}
            CẢNH BÁO HÀNG GIẢ
        {% endif %}
    </div>

    <!-- MESSAGE -->
    <div class="message">
        {{ message }}
    </div>

    {% if product %}
        <div class="divider"></div>

        <div class="product">
            {% if product.image %}
                <img src="{{ product.image }}" alt="Ảnh sản phẩm">
            {% endif %}

            <h3>{{ product.name }}</h3>

            {% if product.brand %}
                <div class="meta">Thương hiệu: {{ product.brand }}</div>
            {% endif %}

            {% if product.description %}
                <div class="meta">Mô tả: {{ product.description }}</div>
            {% endif %}

            <div class="badge
                {% if status == 'AUTHENTIC' %}green
                {% elif status == 'SCANNED' %}yellow
                {% else %}red{% endif %}
            ">
                🔍 Số lần quét: {{ scan_count }}
            </div>
        </div>
    {% endif %}

    <div class="footer">
        © {{ now|default:"2025" }} • Hệ thống xác minh sản phẩm
    </div>
</div>

</body>
</html>
"""
