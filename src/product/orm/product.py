from io import BytesIO
from uuid import UUID

import openpyxl
from cloudinary.uploader import upload
from django.db import transaction
from django.db.models import Prefetch, Q
from django.http import HttpResponse
from openpyxl.utils import get_column_letter

from product.models import Brand, Product, ProductImage, Review
from product.schemas import ProductRequestSchema, SearchFilterSortSchema


class ProductORM:
    @staticmethod
    def get_product_file():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sách sản phẩm"

        headers = [
            "Tên sản phẩm",
            "Mã sản phẩm",
            "Giá gốc",
            "Giá bán",
            "Thương hiệu",
            "Phân loại",
            "Mô tả",
            "Số lượng trong kho",
        ]
        ws.append(headers)

        for i, _ in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 25

        output = BytesIO()

        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=product_list.xlsx"
        return response

    @staticmethod
    @transaction.atomic
    def create(payload: ProductRequestSchema) -> Product:
        product_brand, _ = Brand.objects.get_or_create(name=payload.brand_name)
        product = Product(**payload.dict(exclude={"brand", "brand_name"}))
        product.brand = product_brand
        product.save()
        return product

    @staticmethod
    @transaction.atomic
    def create_multiple(product_file):
        workbook = openpyxl.load_workbook(product_file)
        sheet = workbook.active

        products = []
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            (
                name,
                code,
                origin_price,
                sale_price,
                brand_name,
                type_,
                description,
                quantity_in_stock,
            ) = row

            if not all(
                [name, code, origin_price, sale_price, type_, quantity_in_stock]
            ):
                continue

            brand = None
            if brand_name:
                brand, _ = Brand.objects.get_or_create(name=brand_name)

            product, _ = Product.objects.update_or_create(
                code=code,
                defaults={
                    "name": name,
                    "origin_price": origin_price,
                    "sale_price": sale_price,
                    "brand": brand,
                    "type": type_,
                    "description": description,
                    "quantity_in_stock": quantity_in_stock,
                },
            )

            products.append(product)

        return products

    @staticmethod
    @transaction.atomic
    def upload_image(product: Product, image_files: list):
        product_images = []
        for idx, image_file in enumerate(image_files):
            result = upload(
                image_file.file,
                folder="product_images/",
                public_id=f"product_{product.uid}_{idx}",
                overwrite=True,
            )

            product_image = ProductImage.objects.create(
                product=product,
                url=result["secure_url"],
                is_main=(idx == 0),
            )
            product_images.append(product_image)

        return product_images

    @staticmethod
    def get_all(payload: SearchFilterSortSchema):
        query = Q(deleted=False)
        if payload.search:
            query &= Q(name__icontains=payload.search)
        if payload.brand:
            query &= Q(brand__uid=payload.brand)
        if payload.min_price is not None:
            query &= Q(sale_price__gte=payload.min_price)
        if payload.max_price is not None:
            query &= Q(sale_price__lte=payload.max_price)
        sort_order = "" if payload.sort == "asc" else "-"
        order_by_field = f"{sort_order}sale_price"
        return (
            Product.objects.filter(query)
            .select_related("brand")
            .prefetch_related(
                Prefetch(
                    "image_fk_product",
                    queryset=ProductImage.objects.all(),
                    to_attr="images",
                )
            )
            .order_by(order_by_field)
        )

    @staticmethod
    def get_by_uid(uid: UUID):
        return (
            Product.objects.filter(uid=uid, deleted=False)
            .select_related("brand")
            .prefetch_related(
                Prefetch(
                    "image_fk_product",
                    queryset=ProductImage.objects.all(),
                    to_attr="images",
                ),
                Prefetch(
                    "review_fk_product",
                    queryset=Review.objects.select_related("user"),
                    to_attr="reviews",
                ),
            )
            .first()
        )

    @staticmethod
    def update(product: Product, payload: ProductRequestSchema):
        data = payload.dict(exclude_unset=True, exclude={"brand", "brand_name"})
        for key, value in data.items():
            setattr(product, key, value)

        product_brand = Brand.objects.get(name=payload.brand_name)
        product.brand = product_brand
        product.save()
        return product

    @staticmethod
    def on_off(product: Product):
        product.deleted = not (product.deleted)
        product.save()
        return product

    @staticmethod
    def delete_product(product: Product):
        product.delete()
        return True
